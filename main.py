from pathlib import Path
import utils
import db, db3
import pandas as pd
import calendar
from datetime import datetime
from sqlalchemy.sql import text
import configparser
import os, shutil, pathlib, fnmatch
import numpy as np
from openpyxl import Workbook, load_workbook

##Configs
config = configparser.ConfigParser()
config.read('config.ini')


def cargarPrevired_a_PROPIA():
    recaudacion = input("-> Ingrese el Periodo de recaudación (YYYYMM): ")
    
    db_info_SANNA = {
        "host": config['DEFAULT']['DB_HOST'],
        "port": config['DEFAULT']['DB_PORT'],
        "username": config['DEFAULT']['DB_USER'],
        "password": config['DEFAULT']['DB_PASSWORD'],
        "database": "dgio_SANNA",
    }
    cnn = db3.DatabaseConnection("sqlserver", db_info_SANNA, trust_connection=True)    

    home = Path.cwd()
    ruta = Path(home, 'Archivos', recaudacion, 'PREVIRED')
    files = ruta.glob("*.txt")
    
    for file in (files):
        charset = utils.get_charset(file)
        df_plano = pd.read_csv(file, encoding=charset, sep='|')
        df_plano.to_sql(file.name, cnn.engine, if_exists='replace', index=False)
    
def generarSalidaEnteraSANNA_Recaudacion():
    recaudacion = input("-> Ingrese el Periodo de recaudación (YYYYMM): ")
    
    db_info_reca = {
        "host": config['DEFAULT']['DB_HOST'],
        "port": config['DEFAULT']['DB_PORT'],
        "username": config['DEFAULT']['DB_USER'],
        "password": config['DEFAULT']['DB_PASSWORD'],
        "database": "DW_recaudacion_previred",
    }
    db_info_SANNA = {
        "host": config['DEFAULT']['DB_HOST'],
        "port": config['DEFAULT']['DB_PORT'],
        "username": config['DEFAULT']['DB_USER'],
        "password": config['DEFAULT']['DB_PASSWORD'],
        "database": "dgio_SANNA",
    }
    
    # create substring using slice
    año = recaudacion[0:4]
    mes = recaudacion[4:6]
    res = calendar.monthrange(int(año), int(mes))
    dia = res[1]
    last_day = ""+str(año)+str(mes)+str(dia)
    now = datetime.now()
    dt_string =  now.strftime("%Y-%m-%d %H:%M")
    print('--------------------------------------------------------------------------------')
    print(f'|{dt_string}| -> Generando Sabana Plano 1 -> {recaudacion} ')
    
    cnn = db3.DatabaseConnection("sqlserver", db_info_reca, trust_connection=True)   
    df_sabana_plano_1 = pd.read_sql_query('''
                                            select *, '1' as tipo_declaracion
                                            from FactEncabezado fe
                                            left join FactAntecedentes fa on fa.numero_folio = fe.numero_folio and fa.subgrupo <>'DNP' and fa.periodo_remuneracion >= '2017-04-01'
                                            where (fe.recaudacion = '''+recaudacion+'''
                                            and FORMAT(fa.periodo_remuneracion, 'yyyyMM') >= 201704
                                            and fe.tipo_planilla in (0,2,3))
                                            union all
                                            select *, '2' as tipo_declaracion
                                            from FactEncabezado fe
                                            left join FactAntecedentes fa on fa.numero_folio = fe.numero_folio and fa.subgrupo = 'DNP' 
                                            where fe.recaudacion = '''+recaudacion+'''
                                            and fe.tipo_planilla = 1
                                            and fa.numero_folio not in (select numero_folio from FactEncabezado fe2 where fe2.recaudacion = fe.recaudacion and fe2.tipo_planilla in (0,2,3))
                                            ''', cnn.engine)
    
    now = datetime.now()
    dt_string =  now.strftime("%Y-%m-%d %H:%M")
    print(f'|{dt_string}| -> Finalizado Generacion Sabana Plano 1 -> {recaudacion}')
    print('--------------------------------------------------------------------------------')
    
    now = datetime.now()
    dt_string =  now.strftime("%Y-%m-%d %H:%M")
    print('--------------------------------------------------------------------------------')
    print(f'|{dt_string}| -> Cargando Sabana Plano 1 -> {recaudacion}')
    cnn2 = db3.DatabaseConnection("sqlserver", db_info_SANNA, trust_connection=True)
    df_sabana_plano_1.to_sql('SABANA_'+recaudacion+'_PLANO_1', cnn2.engine, if_exists='replace', index=False, chunksize=100000)
    print(f'|{dt_string}| -> Finalizado Carga Sabana Plano 1 -> {recaudacion}')
    print('--------------------------------------------------------------------------------')
    
def generarSANNA_Recaudacion():
    recaudacion = input("-> Ingrese el Periodo de recaudación (YYYYMM): ")
    cuotas = input("-> Ingrese el N° Cuota de TGR (1-2-1R): ")
    list_cuotas = cuotas.split('-')
    año_cuota = input("-> Ingrese el Año de la cuota TGR: ")

    db_info_SANNA = {
        "host": config['DEFAULT']['DB_HOST'],
        "port": config['DEFAULT']['DB_PORT'],
        "username": config['DEFAULT']['DB_USER'],
        "password": config['DEFAULT']['DB_PASSWORD'],
        "database": "dgio_SANNA",
    }
    cnn = db3.DatabaseConnection("sqlserver", db_info_SANNA, trust_connection=True)
    
    cnn.execute_stored_procedure_nreturn("[dbo].[SP_SANNA_Crear_Planos] @recaudacion='"+recaudacion+"'")
    
    now = datetime.now()
    dt_string =  now.strftime("%Y-%m-%d %H:%M")
    print('--------------------------------------------------------------------------------')
    print(f'|{dt_string}| -> Cargando Previred Plano 1 -> {recaudacion}')
    cnn.execute_stored_procedure_nreturn("[dbo].[SP_SANNA_Cargar_P1_Previred] @recaudacion='"+recaudacion+"'")
    now = datetime.now()
    dt_string =  now.strftime("%Y-%m-%d %H:%M")
    print('--------------------------------------------------------------------------------')
    print(f'|{dt_string}| -> Cargando Previred Plano 2 -> {recaudacion}')
    cnn.execute_stored_procedure_nreturn("[dbo].[SP_SANNA_Cargar_P2_Previred] @recaudacion='"+recaudacion+"'")
    now = datetime.now()  
    
    #Limpiar una sola vez las TGR Insertadas
    cnn.execute_stored_procedure_nreturn("[dbo].[SP_SANNA_LIMPIAR_TABLA_CONSOLIDADA_TGR]  @recaudacion='"+recaudacion+"'")
    now = datetime.now()
        
    for cuota in list_cuotas:
        dt_string =  now.strftime("%Y-%m-%d %H:%M")
        print('--------------------------------------------------------------------------------')
        print(f'|{dt_string}| -> Cargando TGR Plano 1 -> {recaudacion} | Cuota -> {cuota}')
        cnn.execute_stored_procedure_nreturn("[dbo].[SP_SANNA_Cargar_P1_TGR] @recaudacion='"+recaudacion+"', @cuota='"+cuota+"', @cuota_año='"+año_cuota+"'")
        now = datetime.now()
        dt_string =  now.strftime("%Y-%m-%d %H:%M")
        print('--------------------------------------------------------------------------------')
        print(f'|{dt_string}| -> Cargando TGR Plano 2 -> {recaudacion} | Cuota -> {cuota}')
        cnn.execute_stored_procedure_nreturn("[dbo].[SP_SANNA_Cargar_P2_TGR] @recaudacion='"+recaudacion+"', @cuota='"+cuota+"', @cuota_año='"+año_cuota+"'")
        now = datetime.now()
        
    dt_string =  now.strftime("%Y-%m-%d %H:%M")
    print('--------------------------------------------------------------------------------')
    print(f'|{dt_string}| -> Cargando Manuales -> {recaudacion}')
    cnn.execute_stored_procedure_nreturn("[dbo].[SP_SANNA_Cargar_P1_MANUAL] @recaudacion='"+recaudacion+"'")
    
    #cnn.execute_query(f"exec SP_SEGUIMIENTO_SANNA '{recaudacion}'")
    #cnn.execute_query(f"exec SP_XML_SANNA '{recaudacion}'")
def generar_planos_salida():
    recaudacion = input("-> Ingrese el Periodo de recaudación (YYYYMM): ")
    
    db_info_SANNA = {
        "host": config['DEFAULT']['DB_HOST'],
        "port": config['DEFAULT']['DB_PORT'],
        "username": config['DEFAULT']['DB_USER'],
        "password": config['DEFAULT']['DB_PASSWORD'],
        "database": "dgio_SANNA",
    }
    cnn = db3.DatabaseConnection("sqlserver", db_info_SANNA, trust_connection=True)   
    work = Path.cwd()
    ruta_templates = Path(work, 'templates')
    Path(work,'Archivos_salida',recaudacion).mkdir(parents=True, exist_ok=True)
    ruta = Path(work,'Archivos_salida',recaudacion)
    
    
    now = datetime.now()
    dt_string =  now.strftime("%Y-%m-%d %H:%M")
    print('--------------------------------------------------------------------------------')
    print(f'|{dt_string}| -> Generando Plano 1 -> {recaudacion}')
    df_plano_1 = pd.read_sql_query("exec [dbo].[SP_SALIDA_PLANO_1] @recaudacion='"+recaudacion+"'", cnn.engine)    
    archivo = Path(ruta, '30200_PLANO1_'+recaudacion+'.csv')
    df_plano_1.fillna("",inplace=True)
    df_plano_1.to_csv(archivo, sep='|', encoding='UTF-8', index=False, header=False)   

    
    now = datetime.now()
    dt_string =  now.strftime("%Y-%m-%d %H:%M")
    print('--------------------------------------------------------------------------------')
    print(f'|{dt_string}| -> Generando Plano 2 -> {recaudacion}')
    df_plano_2 = pd.read_sql_query("exec [dbo].[SP_SALIDA_PLANO_2] @recaudacion='"+recaudacion+"'", cnn.engine)    
    archivo2 = Path(ruta, '30200_PLANO2_'+recaudacion+'.csv')
    df_plano_2.fillna("",inplace=True)
    df_plano_2.to_csv(archivo2, sep='|', encoding='UTF-8', index=False, header=False)
    
    
    now = datetime.now()
    dt_string =  now.strftime("%Y-%m-%d %H:%M")
    print('--------------------------------------------------------------------------------')
    print(f'|{dt_string}| -> Generando Plano 3 y 4 -> {recaudacion}')
    df_plano_3 = pd.DataFrame()
    archivo3 = Path(ruta, '30200_PLANO3_'+recaudacion+'.csv')
    archivo4 = Path(ruta, '30200_PLANO4_'+recaudacion+'.csv')
    df_plano_3.to_csv(archivo3, sep='|', encoding='UTF-8', index=False)
    df_plano_3.to_csv(archivo4, sep='|', encoding='UTF-8', index=False)
    
    
    now = datetime.now()
    dt_string =  now.strftime("%Y-%m-%d %H:%M")
    print('--------------------------------------------------------------------------------')
    print(f'|{dt_string}| -> Generando XML -> {recaudacion}')
    df_xml = pd.read_sql_query("exec [dbo].[SP_XML_SANNA] @recaudacion='"+recaudacion+"'", cnn.engine)    
    archivo_xml = Path(ruta, '30200_SANNA_RECAUDACION_'+recaudacion+'.xml')
    comilla = " "
    df_xml.to_csv(archivo_xml, encoding='UTF-8', header=False, index=False)
    
    
    #Generar SEGUIMIENTO
    now = datetime.now()
    dt_string =  now.strftime("%Y-%m-%d %H:%M")
    print('--------------------------------------------------------------------------------')
    print(f'|{dt_string}| -> Generando Seguimiento -> {recaudacion}')
    df_seguimiento = pd.read_sql_query("exec [dbo].[SP_SEGUIMIENTO_SANNA] @recaudacion='"+recaudacion+"'", cnn.engine)    
    
    template_seguimiento = Path(ruta_templates, 'Seguimiento_Base.xlsx')
    archivo_seguimiento = Path(ruta, 'SEGUIMIENTO_'+recaudacion+'.xlsx')
    wb = load_workbook(template_seguimiento)
    ws = wb.active
    
    # Iteración por filas del DataFrame:
    for indice_fila, fila in df_seguimiento.iterrows():
        index = int(str(indice_fila))+3
        ws['A'+str(index)] = str(fila['origen'])
        ws['B'+str(index)] = str(fila['periodo'])
        ws['C'+str(index)] = fila['A']
        ws['D'+str(index)] = fila['A1']
        ws['E'+str(index)] = fila['A2']
        ws['F'+str(index)] = fila['A3']
        ws['G'+str(index)] = fila['A4']
        ws['H'+str(index)] = fila['A5']
        ws['I'+str(index)] = fila['A6']
        ws['J'+str(index)] = fila['A7']
        ws['K'+str(index)] = fila['A8']
        ws['L'+str(index)] = fila['B']
        ws['M'+str(index)] = fila['B1']
        ws['N'+str(index)] = fila['B2']
        ws['O'+str(index)] = fila['B3']
        ws['P'+str(index)] = fila['B4']
        ws['Q'+str(index)] = fila['B5']
        ws['R'+str(index)] = fila['B6']
        ws['S'+str(index)] = fila['trabajadores_unicos']
        ws['T'+str(index)] = fila['empleadores_unicos']
        ws['U'+str(index)] = fila['planillas_unicas']
    wb.save(archivo_seguimiento)
    

def generar_planos_porcentual():
    recaudacion = input("-> Ingrese el Periodo de recaudación (YYYYMM): ")
    
    db_info_SANNA = {
        "host": config['DEFAULT']['DB_HOST'],
        "port": config['DEFAULT']['DB_PORT'],
        "username": config['DEFAULT']['DB_USER'],
        "password": config['DEFAULT']['DB_PASSWORD'],
        "database": "dgio_SANNA",
    }
    cnn = db3.DatabaseConnection("sqlserver", db_info_SANNA, trust_connection=True)   
    work = Path.cwd()
    ruta_templates = Path(work, 'templates')
    Path(work,'Archivos_salida',str(recaudacion)+"_Porcentual").mkdir(parents=True, exist_ok=True)
    ruta = Path(work,'Archivos_salida',str(recaudacion)+"_Porcentual")
    
    
    now = datetime.now()
    dt_string =  now.strftime("%Y-%m-%d %H:%M")
    print('--------------------------------------------------------------------------------')
    print(f'|{dt_string}| -> Generando Plano 1 -> {recaudacion}')
    df_plano_1 = pd.read_sql_query("exec [dbo].[SP_SALIDA_PLANO_1_PORCENTUAL] @recaudacion='"+recaudacion+"'", cnn.engine)    
    archivo = Path(ruta, '30200_PLANO1_'+recaudacion+'.csv')
    df_plano_1.fillna("",inplace=True)
    df_plano_1.to_csv(archivo, sep='|', encoding='UTF-8', index=False, header=False)   

    
    now = datetime.now()
    dt_string =  now.strftime("%Y-%m-%d %H:%M")
    print('--------------------------------------------------------------------------------')
    print(f'|{dt_string}| -> Generando Plano 2 -> {recaudacion}')
    df_plano_2 = pd.read_sql_query("exec [dbo].[SP_SALIDA_PLANO_2] @recaudacion='"+recaudacion+"'", cnn.engine)    
    archivo2 = Path(ruta, '30200_PLANO2_'+recaudacion+'.csv')
    df_plano_2.fillna("",inplace=True)
    df_plano_2.to_csv(archivo2, sep='|', encoding='UTF-8', index=False, header=False)
    
    
    now = datetime.now()
    dt_string =  now.strftime("%Y-%m-%d %H:%M")
    print('--------------------------------------------------------------------------------')
    print(f'|{dt_string}| -> Generando Plano 3 y 4 -> {recaudacion}')
    df_plano_3 = pd.DataFrame()
    archivo3 = Path(ruta, '30200_PLANO3_'+recaudacion+'.csv')
    archivo4 = Path(ruta, '30200_PLANO4_'+recaudacion+'.csv')
    df_plano_3.to_csv(archivo3, sep='|', encoding='UTF-8', index=False)
    df_plano_3.to_csv(archivo4, sep='|', encoding='UTF-8', index=False)
    
    
    now = datetime.now()
    dt_string =  now.strftime("%Y-%m-%d %H:%M")
    print('--------------------------------------------------------------------------------')
    print(f'|{dt_string}| -> Generando XML -> {recaudacion}')
    df_xml = pd.read_sql_query("exec [dbo].[SP_XML_SANNA_PORCENTUAL] @recaudacion='"+recaudacion+"'", cnn.engine)    
    archivo_xml = Path(ruta, '30200_SANNA_RECAUDACION_'+recaudacion+'.xml')
    comilla = " "
    df_xml.to_csv(archivo_xml, encoding='UTF-8', header=False, index=False)
    
    
    #Generar SEGUIMIENTO
    now = datetime.now()
    dt_string =  now.strftime("%Y-%m-%d %H:%M")
    print('--------------------------------------------------------------------------------')
    print(f'|{dt_string}| -> Generando Seguimiento -> {recaudacion}')
    df_seguimiento = pd.read_sql_query("exec [dbo].[SP_SEGUIMIENTO_SANNA_PORCENTUAL] @recaudacion='"+recaudacion+"'", cnn.engine)    
    
    template_seguimiento = Path(ruta_templates, 'Seguimiento_Base.xlsx')
    archivo_seguimiento = Path(ruta, 'SEGUIMIENTO_'+recaudacion+'.xlsx')
    wb = load_workbook(template_seguimiento)
    ws = wb.active
    
    # Iteración por filas del DataFrame:
    for indice_fila, fila in df_seguimiento.iterrows():
        index = int(str(indice_fila))+3
        ws['A'+str(index)] = str(fila['origen'])
        ws['B'+str(index)] = str(fila['periodo'])
        ws['C'+str(index)] = fila['A']
        ws['D'+str(index)] = fila['A1']
        ws['E'+str(index)] = fila['A2']
        ws['F'+str(index)] = fila['A3']
        ws['G'+str(index)] = fila['A4']
        ws['H'+str(index)] = fila['A5']
        ws['I'+str(index)] = fila['A6']
        ws['J'+str(index)] = fila['A7']
        ws['K'+str(index)] = fila['A8']
        ws['L'+str(index)] = fila['B']
        ws['M'+str(index)] = fila['B1']
        ws['N'+str(index)] = fila['B2']
        ws['O'+str(index)] = fila['B3']
        ws['P'+str(index)] = fila['B4']
        ws['Q'+str(index)] = fila['B5']
        ws['R'+str(index)] = fila['B6']
        ws['S'+str(index)] = fila['trabajadores_unicos']
        ws['T'+str(index)] = fila['empleadores_unicos']
        ws['U'+str(index)] = fila['planillas_unicas']
    wb.save(archivo_seguimiento)
    


def cargar_gravamenes():
    #pd.options.display.float_format = '{:,.2f}'.format
    #input filename csv file
    periodo = input('Ingrese el periodo a cargar (YYYYmm):')
    #try catch
    try:
        db_gravamenes = {
            "host": config['DEFAULT']['DB_HOST'],
            "port": config['DEFAULT']['DB_PORT'],
            "username": config['DEFAULT']['DB_USER'],
            "password": config['DEFAULT']['DB_PASSWORD'],
            "database": "GRAVAMENES",
        }
        #read csv file all fields as string
        df = pd.read_csv(f'GRAVAMENES\\g_{periodo}.csv', sep=';',dtype=str)
        #insert the values in the database
        cnn = db3.DatabaseConnection("sqlserver", db_gravamenes, trust_connection=True)   
        sql = f"DELETE FROM consolidado_gravamen WHERE periodo = {periodo};"
        cnn.execute_query_nreturn(text(sql))
        
        for index, row in df.iterrows():         
            sql = f"insert into consolidado_gravamen (anio_devengado,mes_devengado,anio_pago,mes_pago,dia_pago,tasa_interes,reajuste,periodo) values ('{row['anio_devengado']}','{row['mes_devengado']}',{row['anio_pago']},{row['mes_pago']},{row['dia_pago']},{row['tasa_interes'].replace(',','.')},{row['reajuste'].replace(',','.')}, '{row['periodo']}')"
            #print(sql)
            cnn.execute_query_nreturn(text(sql))
        
        print("*GRAVAMENES CARGADOS*")
        sql = f" update consolidado_gravamen set fecha_pago_date = CAST(CONCAT(anio_pago, '-',FORMAT( mes_pago, '00'), '-', FORMAT( dia_pago, '00')) as date), periodo_deuda_int = CAST(CONCAT(anio_devengado,FORMAT( mes_devengado, '00')) as int)  where periodo = {periodo}"
        cnn.execute_query_nreturn(text(sql))
        print("*GRAVAMENES ACTUALIZADOS*")
        
    except Exception as e: 
        print(e)
        print('**** error al cargar gravamenes ***')
# check if is main
if __name__ == '__main__':
    utils.borra_pantalla()
    while True:
        print("""
        MENU PRINCIPAL
        ==============
        1. Generar SANNA Recaudacion Propia 
        2. Generar Planos SANNA
        3. Cargar Gravamenes
        4. Generar Planos SANNA - Porcentual
        0. Salir
        """)
        opcion = input("Ingrese una opción: ")
        if opcion == "1":
            generarSANNA_Recaudacion()
        elif opcion == "2":
            generar_planos_salida()
        elif opcion == "3":
            cargar_gravamenes()
        elif opcion == "4":    
            generar_planos_porcentual()
        elif opcion == "0":
            print("Saliendo...")
            break
